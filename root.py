import time
from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook

api = "enter your api"

def get_data_for_n(n):
    url = f"https://{api}/ enter url  {n} "
    response = requests.get(url)
    if response.status_code == 200:
        html_content = response.content
        soup = BeautifulSoup(html_content, 'html.parser')
        data = {}
        rows = soup.find_all('tr')
        for row in rows:
            columns = row.find_all('th')
            if len(columns) == 2:
                key = columns[0].get_text(strip=True)
                value = columns[1].get_text(strip=True)
                data[key] = value
        return data
    else:
        print(f"Failed to fetch data for n = {n}")
        return None

def main(start, end):
    all_data = []
    n_values = range(start, end + 1)
    for n in n_values:
        data = get_data_for_n(n)
        if data:
            all_data.append(data)
        print(f"Processed file with n={n}")
    save_to_excel(all_data, "all_data_name2.xlsx")


def save_to_excel(data_list, filename):
    if not data_list:
        print("No data to save.")
        return

    wb = Workbook()
    ws = wb.active

    # Writing headers
    headers = data_list[0].keys()
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # Writing data
    for row, data in enumerate(data_list, start=2):
        for col, value in enumerate(data.values(), start=1):
            ws.cell(row=row, column=col, value=value)

    wb.save(filename)
    print(f"All data saved to {filename}")

if __name__ == "__main__":
    start =1000
    end = 5000
    main(start, end)
